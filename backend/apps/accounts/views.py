from rest_framework import status, viewsets, permissions
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import JSONParser
from django.contrib.auth import logout
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from openpyxl import Workbook
from io import BytesIO
from django.http import HttpResponse
from django.db.models import Q

from .models import User, SMSVerificationCode
from .serializers import (
    UserSerializer,
    UserCreateSerializer,
    UserUpdateSerializer,
    LoginSerializer,
    TokenSerializer,
    SendSMSVerificationSerializer,
    VerifySMSSerializer,
)
from .permissions import IsAdminOrReadOnly, IsAdmin
from .sms_service import sms_service
from django.conf import settings
import logging

logger = logging.getLogger(__name__)


class LoginView(APIView):
    """Login endpoint"""
    permission_classes = [permissions.AllowAny]
    parser_classes = [JSONParser]
    
    def post(self, request):
        """Login and get JWT tokens"""
        # Debug logging
        import logging
        logger = logging.getLogger(__name__)
        
        # DRF should parse JSON automatically via request.data
        data = request.data
        logger.info(f"Login attempt - data type: {type(data)}, data: {data}, content_type: {request.content_type}")
        
        # Ensure data is a dict
        if not isinstance(data, dict):
            logger.error(f"Data is not a dict: {data}, type: {type(data)}")
            return Response(
                {'non_field_errors': ['Invalid request format.']},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = LoginSerializer(data=data, context={'request': request})
        if not serializer.is_valid():
            logger.error(f"Serializer errors: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        user = serializer.validated_data['user']
        tokens = TokenSerializer.get_tokens_for_user(user)
        logger.info(f"Login successful for user: {user.phone}")
        return Response(tokens, status=status.HTTP_200_OK)


class RegisterView(APIView):
    """Register endpoint"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        """Register new user"""
        serializer = UserCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        
        # Mark user as verified after successful registration with SMS code
        if request.data.get('verification_code'):
            user.verified = True
            user.save()
        
        tokens = TokenSerializer.get_tokens_for_user(user)
        return Response(tokens, status=status.HTTP_201_CREATED)


class LogoutView(APIView):
    """Logout endpoint"""
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        """Logout user"""
        try:
            logout(request)
        except Exception:
            pass
        return Response({'message': 'Successfully logged out'}, status=status.HTTP_200_OK)


class MeView(APIView):
    """Get current user endpoint"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        """Get current user"""
        serializer = UserSerializer(request.user)
        return Response(serializer.data)


class UserViewSet(viewsets.ModelViewSet):
    """User management ViewSet"""
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAdmin]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['role', 'verified', 'is_active']
    search_fields = ['phone', 'email', 'full_name', 'iin']
    ordering_fields = ['created_at', 'full_name']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return UserUpdateSerializer
        return UserSerializer
    
    def create(self, request, *args, **kwargs):
        """Override create to return generated password if applicable"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        
        # Если пароль был сгенерирован, возвращаем его в ответе
        response_data = UserSerializer(user).data
        if hasattr(user, '_generated_password'):
            response_data['generated_password'] = user._generated_password
        
        headers = self.get_success_headers(serializer.data)
        return Response(response_data, status=status.HTTP_201_CREATED, headers=headers)
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export users to Excel"""
        users = self.filter_queryset(self.get_queryset())
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Headers
        headers = ['ID', 'Phone', 'Full Name', 'Email', 'IIN', 'Role', 'Verified', 'City', 'Organization', 'Created At']
        ws.append(headers)
        
        # Data
        for user in users:
            ws.append([
                user.id,
                user.phone,
                user.full_name or '',
                user.email or '',
                user.iin or '',
                user.get_role_display(),
                'Yes' if user.verified else 'No',
                user.city or '',
                user.organization or '',
                user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else '',
            ])
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=users.xlsx'
        return response
    
    @action(detail=False, methods=['post'])
    def import_users(self, request):
        """Import users from Excel"""
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        file = request.FILES['file']
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            
            imported = 0
            errors = []
            
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    phone = str(row[1]) if row[1] else None
                    if not phone:
                        continue
                    
                    # Check if user exists
                    if User.objects.filter(phone=phone).exists():
                        errors.append({'row': row, 'error': f'User with phone {phone} already exists'})
                        continue
                    
                    User.objects.create_user(
                        phone=phone,
                        password='default123',  # Should be changed by user
                        full_name=row[2] or '',
                        email=row[3] or None,
                        iin=str(row[4]) if row[4] else None,
                        role=row[5] or 'student',
                        city=row[7] or '',
                        organization=row[8] or '',
                    )
                    imported += 1
                except Exception as e:
                    errors.append({'row': row, 'error': str(e)})
            
            return Response({
                'imported': imported,
                'errors': errors
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class SendSMSVerificationView(APIView):
    """Send SMS verification code"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        """Send SMS verification code"""
        serializer = SendSMSVerificationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        phone = serializer.validated_data['phone']
        purpose = serializer.validated_data['purpose']
        
        # Normalize phone number
        normalized_phone = ''.join(filter(str.isdigit, str(phone)))
        if normalized_phone.startswith('8'):
            normalized_phone = '7' + normalized_phone[1:]
        if not normalized_phone.startswith('7'):
            normalized_phone = '7' + normalized_phone
        
        try:
            # Generate verification code
            verification_code = SMSVerificationCode.generate_code(normalized_phone, purpose)
            
            # Send SMS via SMSC.kz
            sms_result = sms_service.send_verification_code(
                normalized_phone,
                verification_code.code,
                purpose
            )
            
            if not sms_result['success']:
                logger.error(f"Failed to send SMS to {normalized_phone}: {sms_result.get('error')}")
                return Response(
                    {
                        'error': sms_result.get('error', 'Failed to send SMS'),
                        'message': sms_result.get('message', 'SMS sending failed')
                    },
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            response_data = {
                'message': 'SMS verification code sent successfully',
                'expires_at': verification_code.expires_at.isoformat(),
            }
            
            # In development mode (when SMSC credentials not configured), return OTP code
            is_smsc_configured = (
                hasattr(settings, 'SMSC_LOGIN') and 
                settings.SMSC_LOGIN and 
                hasattr(settings, 'SMSC_PASSWORD') and 
                settings.SMSC_PASSWORD
            )
            
            if not is_smsc_configured:
                response_data['otp_code'] = verification_code.code
                response_data['debug'] = True
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error sending SMS verification code: {str(e)}")
            return Response(
                {'error': 'Failed to send verification code', 'detail': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class VerifySMSView(APIView):
    """Verify SMS code"""
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        """Verify SMS code"""
        serializer = VerifySMSSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        phone = serializer.validated_data['phone']
        code = serializer.validated_data['code']
        purpose = serializer.validated_data['purpose']
        
        # Normalize phone number
        normalized_phone = ''.join(filter(str.isdigit, str(phone)))
        if normalized_phone.startswith('8'):
            normalized_phone = '7' + normalized_phone[1:]
        if not normalized_phone.startswith('7'):
            normalized_phone = '7' + normalized_phone
        
        try:
            # Find the most recent unverified code for this phone and purpose
            verification_code = SMSVerificationCode.objects.filter(
                phone=normalized_phone,
                purpose=purpose,
                is_verified=False
            ).order_by('-created_at').first()
            
            if not verification_code:
                return Response(
                    {'verified': False, 'error': 'Verification code not found or already used'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Verify the code
            if verification_code.verify(code):
                return Response(
                    {'verified': True, 'message': 'Code verified successfully'},
                    status=status.HTTP_200_OK
                )
            else:
                return Response(
                    {'verified': False, 'error': 'Invalid or expired code'},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except Exception as e:
            logger.error(f"Error verifying SMS code: {str(e)}")
            return Response(
                {'verified': False, 'error': 'Verification failed', 'detail': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

